"""Step 4 — compare an external solution (EVOX) against VRPP + Lookahead.

    python scripts/04_compare_evox.py --config config/instance_491_C7.yaml \
        --external data/raw/evox_routes_491_C7.yaml

Builds a three-block comparison table:

  * **EVOX reported**      — the KPI as EVOX itself reports them (`reported:` in
                             the external YAML). Blank where not supplied.
  * **VRPP fixing EVOX**   — the very same routes and visiting order, measured
                             with the ORS matrix, the YAML parameters and this
                             project's MustGo / MustGoLA classification.
  * **VRPP optimal**       — the free optimisation from `03_run_vrpp.py`.

Both computed blocks go through the same measuring code (`fixed_routes.py`), so
any difference between them comes from the routes, never from the yardstick.
"""
from __future__ import annotations

import argparse

import pandas as pd
import yaml

from _common import base_parser

from vrpp_lookahead import Config, load_instance
from vrpp_lookahead.fixed_routes import (REPORT_ROWS, TOTAL_ROWS, TOTAL_ROWS_SHIFT,
                                         evaluate_fixed_routes,
                                         routes_from_solution_workbook)

BLANK = ''
NOT_REPORTED = 'n/a'   # the external system does not publish this figure


def _fmt(value):
    return NOT_REPORTED if value is None else value


def _diff(reported, computed):
    """reported - computed, blank when EVOX did not supply the figure."""
    if reported is None or computed is None:
        return BLANK
    return round(float(reported) - float(computed), 4)


def build_table(metrics_fix, metrics_opt, reported: dict) -> pd.DataFrame:
    routes_fix = sorted(metrics_fix.per_route['route'])
    routes_opt = sorted(metrics_opt.per_route['route'])
    fix = metrics_fix.per_route.set_index('route')
    opt = metrics_opt.per_route.set_index('route')
    rep_totals = (reported or {}).get('totals') or {}

    rows = []
    for title, key in REPORT_ROWS:
        row = {'Metric': title}
        for r in routes_fix:
            rep = ((reported or {}).get(r) or {}).get(key)
            row[f'EVOX_rep R{r}'] = _fmt(rep)
            row[f'VRPP_fix R{r}'] = fix.loc[r, key]
            row[f'Diff R{r}'] = _diff(rep, fix.loc[r, key])
        for r in routes_opt:
            row[f'VRPP_opt R{r}'] = opt.loc[r, key]
        rows.append(row)

    for title, key in TOTAL_ROWS + TOTAL_ROWS_SHIFT:
        row = {'Metric': title}
        for r in routes_fix:
            row[f'EVOX_rep R{r}'] = BLANK
            row[f'VRPP_fix R{r}'] = BLANK
            row[f'Diff R{r}'] = BLANK
        for r in routes_opt:
            row[f'VRPP_opt R{r}'] = BLANK
        row['EVOX_rep TOTAL'] = _fmt(rep_totals.get(key))
        row['VRPP_fix TOTAL'] = metrics_fix.totals[key]
        row['VRPP_opt TOTAL'] = metrics_opt.totals[key]
        row['Opt - Fix'] = round(metrics_opt.totals[key] - metrics_fix.totals[key], 4)
        rows.append(row)

    df = pd.DataFrame(rows)

    # per-route metrics also get a TOTAL / Opt-Fix column where summing makes sense
    # `shift_total_h` is deliberately NOT summable: a shift is a per-crew limit, so
    # the meaningful aggregate is the longest route (see TOTAL_ROWS_SHIFT), not the sum.
    summable = {'n_bins', 'n_mustgo', 'n_mustgo_la', 'n_optional', 'weight_kg',
                'distance_km', 'travel_time_min', 'profit_euro',
                'shift_driving_min', 'shift_service_min'}
    for i, (title, key) in enumerate(REPORT_ROWS):
        if key in summable:
            f = float(fix[key].sum())
            o = float(opt[key].sum())
            df.loc[i, 'VRPP_fix TOTAL'] = round(f, 4)
            df.loc[i, 'VRPP_opt TOTAL'] = round(o, 4)
            df.loc[i, 'Opt - Fix'] = round(o - f, 4)
            rep_vals = [((reported or {}).get(r) or {}).get(key) for r in routes_fix]
            if all(v is not None for v in rep_vals):
                df.loc[i, 'EVOX_rep TOTAL'] = round(sum(rep_vals), 4)

    ordered = (['Metric']
               + [c for r in routes_fix for c in (f'EVOX_rep R{r}',)]
               + ['EVOX_rep TOTAL']
               + [c for r in routes_fix for c in (f'VRPP_fix R{r}',)]
               + ['VRPP_fix TOTAL']
               + [c for r in routes_fix for c in (f'Diff R{r}',)]
               + [c for r in routes_opt for c in (f'VRPP_opt R{r}',)]
               + ['VRPP_opt TOTAL', 'Opt - Fix'])
    for c in ordered:
        if c not in df.columns:
            df[c] = BLANK
    return df[ordered].fillna(BLANK)


def solution_provenance(path) -> dict:
    """Which optimisation run the 'VRPP optimal' block was read from."""
    try:
        kpi = pd.read_excel(path, sheet_name='2_KPI_General').set_index('KPI')['Value']
        return {'file': str(path), 'gap_pct': float(kpi['MIP Gap']),
                'solver_time_h': float(kpi['Solver time'].iloc[-1]
                                       if hasattr(kpi['Solver time'], 'iloc')
                                       else kpi['Solver time']),
                'status': int(float(kpi['Status (2=Optimal,9=TimeLimit)']))}
    except Exception:
        return {'file': str(path)}


def build_notes(metrics_fix, metrics_opt, reported: dict, cfg: Config,
                provenance: dict | None = None) -> pd.DataFrame:
    """Caveats a reader needs to interpret the table, derived from the data itself.

    Keeps the workbook self-contained: why blocks differ, which gaps come from a
    different definition rather than a different result, and what simply cannot
    be reconciled because the external system does not publish it.
    """
    mod = cfg.model
    fix = metrics_fix.per_route.set_index('route')
    notes = [{
        'Topic': 'Measuring method',
        'Detail': 'The "VRPP fixing EVOX" and "VRPP optimal" blocks are produced by the '
                  'same code over the same ORS matrix and YAML parameters, so any gap '
                  'between them comes from the routes, not from the yardstick.',
    }]

    for r in sorted(fix.index):
        rep = (reported or {}).get(r) or {}
        w_rep, d_rep, p_rep = rep.get('weight_kg'), rep.get('distance_km'), rep.get('profit_euro')
        if w_rep is not None:
            gap = round(float(w_rep) - float(fix.loc[r, 'weight_kg']), 3)
            notes.append({
                'Topic': f'Route {r} — weight',
                'Detail': (f'External {w_rep} kg vs computed {fix.loc[r, "weight_kg"]} kg '
                           f'(gap {gap:+} kg). '
                           + ('Identical: the bin levels of both systems agree.' if gap == 0 else
                              'A non-zero gap over the same bin ids means either an id in the '
                              'supplied list is not the one actually served, or the reported '
                              'figure is a transcription slip. Not resolvable without the '
                              'external per-bin breakdown, which is not published.')),
            })
        if d_rep is not None:
            gap = round(float(d_rep) - float(fix.loc[r, 'distance_km']), 3)
            pct = gap / float(fix.loc[r, 'distance_km']) * 100
            notes.append({
                'Topic': f'Route {r} — distance',
                'Detail': (f'External {d_rep} km vs computed {fix.loc[r, "distance_km"]} km '
                           f'(gap {gap:+} km, {pct:+.1f} %). Computed over the supplied visiting '
                           f'order with the {cfg.ors.transport_mode} ORS profile.'),
            })
        # Same for the stop time: n_bins and the reported service minutes pin down
        # the rate the external system assumed, which is the other half of the day.
        service_rep = rep.get('shift_service_min')
        n_rep = rep.get('n_bins') or fix.loc[r, 'n_bins']
        if service_rep and n_rep:
            implied = float(service_rep) / float(n_rep)
            same = abs(implied - cfg.shift.service_time_min) < 0.01
            notes.append({
                'Topic': f'Route {r} — implied service time',
                'Detail': (f'{service_rep} min of stop time over {n_rep:g} bins is an implied '
                           f'{implied:.2f} min per bin, against the '
                           f'{cfg.shift.service_time_min:g} min configured for this study'
                           + ('— the same assumption, so the stop times are comparable.' if same else
                              f'. Re-measured at {cfg.shift.service_time_min:g} min the same '
                              f'{n_rep:g} stops cost {fix.loc[r, "shift_service_min"]:.1f} min. '
                              f'That row differs by definition, not by result — and it is the '
                              f'stricter figure that both solutions are held to here.')),
            })
        # Where the external system publishes both a distance and a driving time,
        # the ratio pins down the speed it actually assumed — the one parameter
        # that silently decides whether a route fits in the shift.
        drive_rep = rep.get('shift_driving_min')
        if d_rep is not None and drive_rep:
            implied = float(d_rep) / (float(drive_rep) / 60.0)
            notes.append({
                'Topic': f'Route {r} — implied average speed',
                'Detail': (f'{d_rep} km in {drive_rep} min is an implied {implied:.1f} km/h, '
                           f'against the {cfg.shift.speed_kmh:g} km/h configured for this study'
                           + (' — the same assumption, so the two driving times are comparable.'
                              if abs(implied - cfg.shift.speed_kmh) < 0.5 else
                              f'. At {cfg.shift.speed_kmh:g} km/h the very same kilometres take '
                              f'{fix.loc[r, "shift_driving_min"]:.0f} min, which is why the '
                              f'working-day rows differ even when the route does not.')),
            })
        if p_rep is not None and w_rep is not None and d_rep is not None:
            no_omega = mod.R * float(w_rep) - mod.C * float(d_rep)
            if abs(no_omega - float(p_rep)) < 0.01:
                notes.append({
                    'Topic': f'Route {r} — profit definition',
                    'Detail': (f'External {p_rep} EUR equals R*kg - C*km = {no_omega:.4f} EUR, '
                               f'i.e. it excludes the fixed vehicle cost OMEGA={mod.OMEGA} EUR '
                               f'that this project subtracts. A definition gap, not a result gap.'),
                })

    missing = [k for k in ('n_mustgo', 'n_mustgo_la', 'n_optional', 'cap_used_pct')
               if all(((reported or {}).get(r) or {}).get(k) is None for r in fix.index)]
    if missing:
        notes.append({
            'Topic': 'Rows marked n/a',
            'Detail': ('The external system does not publish these figures, so the cells stay '
                       'empty permanently and no difference can be computed for them: '
                       + ', '.join(missing) + '. The values under "VRPP fixing EVOX" are this '
                       'project\'s classification applied to the external routes.'),
        })

    if provenance:
        status = {2: 'OPTIMAL', 9: 'TIME_LIMIT'}.get(provenance.get('status'), '')
        gap = provenance.get('gap_pct')
        notes.append({
            'Topic': 'Optimal block provenance',
            'Detail': (f'Read from {provenance["file"]}'
                       + (f', which stopped at a proven MIP gap of {gap:.2f} %'
                          if gap is not None else '')
                       + (f' with status {status}' if status else '')
                       + (f' after {provenance["solver_time_h"]:.2f} h'
                          if provenance.get('solver_time_h') is not None else '')
                       + '. A non-zero gap means the true optimum is at most that much better '
                         'than this solution, so every "optimal" figure below is a lower bound '
                         'on the achievable improvement over the external routes.'),
        })

    notes.append({
        'Topic': 'MustGo coverage',
        'Detail': (f'{metrics_fix.totals["n_mustgo_missed"]} MustGo and '
                   f'{metrics_fix.totals["n_mustgo_la_missed"]} MustGo-LookAhead bins are left '
                   f'uncovered by the external routes, against '
                   f'{metrics_opt.totals["n_mustgo_missed"]} and '
                   f'{metrics_opt.totals["n_mustgo_la_missed"]} by the optimal solution. '
                   f'An uncovered MustGo overflows the next day.'),
    })
    return pd.DataFrame(notes)


def build_interpretation(metrics_fix, metrics_opt, reported: dict, cfg: Config,
                         inst) -> pd.DataFrame:
    """The reading of the table: what the numbers mean, in plain sentences."""
    rep_t = (reported or {}).get('totals') or {}
    f, o = metrics_fix.totals, metrics_opt.totals
    fix_route = metrics_fix.per_route
    ext_km = rep_t.get('total_distance_km')
    ext_kg = rep_t.get('total_weight_kg')
    ext_eur = rep_t.get('total_profit_euro')
    ext_ratio = rep_t.get('total_km_per_ton')

    def pct(new, old):
        return f'{(new - old) / old * 100:+.1f} %' if old else 'n/a'

    items = []
    if ext_km and ext_kg:
        more_kg = o['total_weight_kg'] > ext_kg
        more_km = o['total_distance_km'] > ext_km
        if more_kg and not more_km:
            verdict = ('It is not a trade-off between distance and capture: the VRPP wins on '
                       'both at once.')
        elif more_kg and more_km:
            verdict = ('The VRPP buys the extra waste with extra kilometres, so distance alone '
                       'settles nothing here — the km/Ton ratio and the profit line below are '
                       'what decide whether the trade is worth making.')
        elif not more_kg and not more_km:
            verdict = ('The VRPP drives less and collects less: it stops where the next bin no '
                       'longer pays for the kilometres to reach it, which is exactly what the '
                       'objective asks it to do.')
        else:
            verdict = ('The VRPP collects less over more kilometres, which should not happen — '
                       'check the MIP gap on the Notes sheet before reading anything into it.')
        items.append((
            'Headline',
            f'Against the figures EVOX reports for itself, the optimal VRPP collects '
            f'{o["total_weight_kg"]:.1f} kg versus {ext_kg:.1f} kg ({pct(o["total_weight_kg"], ext_kg)}) '
            f'while driving {o["total_distance_km"]:.2f} km versus {ext_km:.1f} km '
            f'({pct(o["total_distance_km"], ext_km)}), with '
            f'{o["n_routes"]} vehicle(s) against {f["n_routes"]}. {verdict}'))
    if ext_eur:
        items.append((
            'Profit',
            f'{o["total_profit_euro"]:.2f} EUR versus {ext_eur:.2f} EUR '
            f'({pct(o["total_profit_euro"], ext_eur)}). Note the two systems define profit '
            f'differently — EVOX omits the fixed vehicle cost OMEGA — but at {cfg.model.OMEGA} EUR '
            f'per vehicle the effect is negligible next to this gap.'))
    if ext_ratio:
        items.append((
            'Efficiency ratio',
            f'{o["total_km_per_ton"]:.2f} km/Ton versus {ext_ratio:.2f} km/Ton '
            f'({pct(o["total_km_per_ton"], ext_ratio)}). This is the metric that summarises the '
            f'difference best, because it normalises distance by what was actually collected.'))

    opt_route = metrics_opt.per_route
    used = fix_route['cap_used_pct'].tolist()
    used_opt = opt_route['cap_used_pct'].tolist()
    idle = len(fix_route) * cfg.model.Q - f['total_weight_kg']
    # Whether an under-filled vehicle is a fault of the routing or of the network
    # depends on how much waste is on the ground at all — say which one it is.
    on_ground = float(inst.bins['Si_kg'].sum())
    if used_opt and max(used_opt) < 70.0:
        fill_reading = (
            f'Neither fills a vehicle, and neither can: the whole network holds {on_ground:.0f} kg '
            f'today, only {on_ground / cfg.model.Q:.1f} times what a single {cfg.model.Q:g} kg '
            f'vehicle carries — and most of it sits in bins too far apart to be worth the '
            f'kilometres. On this instance the truck is not the scarce resource, so a low fill '
            f'rate is not evidence of a bad route.')
    else:
        fill_reading = (
            f'It leaves {idle:.0f} kg of the capacity it did send unused while paying the fixed '
            f'cost and most of the driving. With bins {inst.dist[inst.dist > 0].mean():.1f} km '
            f'apart on average, an extra stop costs little distance and adds weight, so a '
            f'half-empty vehicle is rarely the cheapest way to run the day.')
    items.append((
        'Vehicle fill',
        f'The external solution runs its {len(fix_route)} vehicle(s) at '
        f'{" and ".join(f"{u:.1f} %" for u in used)} of capacity, against '
        f'{" and ".join(f"{u:.1f} %" for u in used_opt)} for the optimal solution. '
        + fill_reading))

    # Which resource actually stops the optimal solution. Saying "fill the truck"
    # when the clock is what binds would point the operator at the wrong lever.
    sh = cfg.shift
    fill_max = max(used_opt) if used_opt else 0.0
    day_max = o['max_shift_total_h']
    if fill_max >= 95.0:
        binding = (f'vehicle capacity — the fullest route is at {fill_max:.1f} % of Q='
                   f'{cfg.model.Q:g} kg. A bigger vehicle would collect more.')
    elif sh.enforce and day_max >= sh.max_shift_h - 0.15:
        binding = (f'the working day — the longest route runs {day_max:.2f} h against the '
                   f'{sh.max_shift_h:g} h limit, while the fullest vehicle is only '
                   f'{fill_max:.1f} % loaded. Capacity is not the scarce resource here; hours '
                   f'are. More waste can only be reached by adding a vehicle (another 7 h of '
                   f'crew time), never by fitting more into the ones already out.')
    else:
        binding = (f'neither capacity ({fill_max:.1f} % of Q) nor the working day '
                   f'({day_max:.2f} h of {sh.max_shift_h:g} h): the solution stops where the '
                   f'next bin stops paying for the kilometres needed to reach it.')
    items.append(('Binding resource', f'What limits the optimal solution is {binding}'))

    items.append((
        'Critical coverage',
        f'The external routes leave {f["n_mustgo_missed"]} MustGo and {f["n_mustgo_la_missed"]} '
        f'MustGo-LookAhead bin(s) uncovered — an uncovered MustGo overflows the next day — while '
        f'collecting {int(fix_route["n_optional"].sum())} optional bins that are not urgent. The '
        f'optimal solution leaves {o["n_mustgo_missed"]} and {o["n_mustgo_la_missed"]} uncovered '
        f'and collects {int(opt_route["n_optional"].sum())} optional bins.'))

    fleet_note = ('sized freely by the model' if cfg.model.MAX_ROUTES_auto
                  else f'capped at MAX_ROUTES={cfg.model.MAX_ROUTES}')
    items.append((
        'Fleet and route balance',
        f'The external solution sends {len(fix_route)} vehicle(s) '
        f'({" and ".join(f"{d:.2f} km" for d in fix_route["distance_km"])}); the optimal one, '
        f'{fleet_note}, sends {len(opt_route)} '
        f'({" and ".join(f"{d:.2f} km" for d in opt_route["distance_km"])}). The model does not '
        f'balance workloads — it maximises joint profit, so unequal routes are a result, not an '
        f'oversight. If an even workload across drivers is a real operational requirement, it has '
        f'to be added to the model; it is not there today.'))

    limits = ' / '.join(f'{h:g} h' for h in sh.report_h)
    extra_stops = int(opt_route['n_bins'].sum() - fix_route['n_bins'].sum())
    extra_service = float(opt_route['shift_service_min'].sum() - fix_route['shift_service_min'].sum())
    if sh.enforce:
        working_day = (
            f'At {sh.speed_kmh:g} km/h with {sh.service_time_min:g} min per bin, the optimal '
            f'routes take {" and ".join(f"{h:.2f} h" for h in opt_route["shift_total_h"])} '
            f'against {" and ".join(f"{h:.2f} h" for h in fix_route["shift_total_h"])} for the '
            f'external ones, reported against {limits}. The {sh.max_shift_h:g} h day is a HARD '
            f'CONSTRAINT of this run, so every figure above is already what fits in a shift: the '
            f'{extra_stops:+d} stops of difference and the {extra_service:+.0f} min of service '
            f'time they carry are paid for inside the day, not on top of it. Nothing here has to '
            f'be discounted afterwards.')
    else:
        working_day = (
            f'At {sh.speed_kmh:g} km/h with {sh.service_time_min:g} min per bin, the optimal '
            f'routes take {" and ".join(f"{h:.2f} h" for h in opt_route["shift_total_h"])} '
            f'against {" and ".join(f"{h:.2f} h" for h in fix_route["shift_total_h"])} for the '
            f'external ones, measured against limits of {limits}. This reverses the reading of '
            f'every row above: the optimal solution wins on distance, weight and profit because '
            f'visiting a bin costs it nothing in the objective, which prices distance and never '
            f'time. The {extra_stops:+d} stops of difference add {extra_service:+.0f} min of '
            f'service time that the model does not see. Set shift.enforce = true to make the '
            f'limit binding and re-solve; the profit advantage will shrink to whatever fits in '
            f'a shift.')
    items.append(('Working day', working_day))

    items.append((
        'Caveat',
        'The middle block re-measures the external routes with this project\'s yardstick, so the '
        'discrepancies against the reported figures (see the Notes sheet) are measurement '
        'differences, not disagreements about the routes themselves.'))

    return pd.DataFrame(items, columns=['Topic', 'Reading'])


def _format_sheet(ws, table: pd.DataFrame) -> None:
    """Readable widths, frozen header and label column, wrapped headers."""
    from openpyxl.styles import Alignment, Font

    ws.freeze_panes = 'B2'
    ws.column_dimensions['A'].width = 40
    for idx in range(2, len(table.columns) + 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = 15
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    ws.row_dimensions[1].height = 32
    for row in ws.iter_rows(min_row=2, min_col=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')


def _format_text_sheet(ws) -> None:
    from openpyxl.styles import Alignment, Font

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 110
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in ws.iter_rows(min_row=2):
        row[0].alignment = Alignment(vertical='top')
        row[-1].alignment = Alignment(wrap_text=True, vertical='top')


def main() -> None:
    p = base_parser(__doc__)
    p.add_argument('--external', required=True,
                   help='YAML with the external routes (and optionally its reported KPI)')
    p.add_argument('--solution', default=None,
                   help='result workbook of the optimal run '
                        '(default: <results>/Day_01/result_Day_01.xlsx)')
    p.add_argument('--out', default=None,
                   help='output Excel (default: <results>/comparison_<label>.xlsx)')
    args = p.parse_args()

    cfg = Config.from_yaml(args.config)
    inst = load_instance(cfg)

    data = yaml.safe_load(open(args.external, encoding='utf-8')) or {}
    external = {int(k): list(v) for k, v in (data.get('routes') or {}).items()}
    assert external, f'No `routes:` block found in {args.external}'
    reported = {(k if k == 'totals' else int(k)): v
                for k, v in (data.get('reported') or {}).items()}

    results = cfg.path('results', create_dir=True)
    solution = args.solution or (results / 'Day_01' / 'result_Day_01.xlsx')
    assert pd.io.common.file_exists(str(solution)), (
        f'Optimal solution not found: {solution}\nRun scripts/03_run_vrpp.py first.')

    print(f'=== STEP 4 — COMPARISON ({cfg.label}) ===')
    print(f'External routes : {args.external}')
    print(f'Optimal solution: {solution}\n')

    metrics_fix = evaluate_fixed_routes(external, inst, cfg, label='EVOX')
    metrics_opt = evaluate_fixed_routes(routes_from_solution_workbook(solution),
                                        inst, cfg, label='VRPP optimal')

    table = build_table(metrics_fix, metrics_opt, reported)
    print(table.to_string(index=False))

    notes = build_notes(metrics_fix, metrics_opt, reported, cfg,
                        solution_provenance(solution))
    print('\n--- Notes ---')
    for _, n in notes.iterrows():
        print(f'  * {n["Topic"]}: {n["Detail"]}')

    interpretation = build_interpretation(metrics_fix, metrics_opt, reported, cfg, inst)
    print('\n--- Interpretation ---')
    for _, n in interpretation.iterrows():
        print(f'  * {n["Topic"]}: {n["Reading"]}')

    out = args.out or (results / f'comparison_{cfg.label}.xlsx')
    with pd.ExcelWriter(out, engine='openpyxl') as w:
        table.to_excel(w, sheet_name='Comparison', index=False)
        interpretation.to_excel(w, sheet_name='Interpretation', index=False)
        notes.to_excel(w, sheet_name='Notes', index=False)
        metrics_fix.per_route.to_excel(w, sheet_name='EVOX_routes_detail', index=False)
        metrics_opt.per_route.to_excel(w, sheet_name='VRPP_optimal_detail', index=False)
        pd.DataFrame({'id_contentor': metrics_fix.not_visited}).to_excel(
            w, sheet_name='EVOX_not_visited', index=False)

        _format_sheet(w.sheets['Comparison'], table)
        _format_text_sheet(w.sheets['Interpretation'])
        _format_text_sheet(w.sheets['Notes'])
    print(f'\nComparison written to: {out}')


if __name__ == '__main__':
    main()
